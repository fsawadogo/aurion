"""Build the consolidated effort estimation workbook.

Merges the old iOS-focused `MVP_Effort_Estimation_Aurion.xlsx` and the
v4.1 web-portal audit into a single source-of-truth file with status
reconciled to as-of 2026-05-22 and a second `Claude-Code-Days` estimate
column scaled per the rates observed in the autonomous workflow.

Run from repo root:
    python3 scripts/build_effort_workbook.py

Writes: MVP_Effort_Estimation_Aurion_v5.xlsx (overwrites if present).
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

OUT_PATH = "MVP_Effort_Estimation_Aurion_v5.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# Status palette (matches the v4.1 emoji legend)
# ─────────────────────────────────────────────────────────────────────────────
STATUS_FILLS = {
    "Done":          PatternFill("solid", fgColor="C6EFCE"),  # 🟢 light green
    "Verify":        PatternFill("solid", fgColor="FFEB9C"),  # 🟡 light amber
    "Not Started":   PatternFill("solid", fgColor="FFD7A8"),  # 🟠 light orange
    "Blocker":       PatternFill("solid", fgColor="FFC7CE"),  # 🔴 light red
    "Future":        PatternFill("solid", fgColor="E4B7E4"),  # 🟣 light purple
    "In Progress":   PatternFill("solid", fgColor="BDD7EE"),  # 🔵 light blue
}
STATUS_EMOJI = {
    "Done":        "🟢",
    "Verify":      "🟡",
    "Not Started": "🟠",
    "Blocker":     "🔴",
    "Future":      "🟣",
    "In Progress": "🔵",
}

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=11)

THIN = Side(style="thin", color="BFBFBF")
ALL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")


# ─────────────────────────────────────────────────────────────────────────────
# Data — every row touched at least once by the audit; status reflects what's
# actually in git as of 2026-05-22.
# ─────────────────────────────────────────────────────────────────────────────

# Columns: ID, Name, Status, Complexity, Human-Days (P50), Claude-Code-Days,
#          Commit / Notes, Pre-requisites
# Lane is added per-sheet.

# ── MOBILE (iOS) — fundamentally complete after the UI sweep + Live Activity ─
MOBILE = [
    # Screens — all shipped
    ("S001", "SplashView", "Done", "S", 1, 0.25, "Built. UI-P1 token refresh applied.", "—"),
    ("S002", "AuthView + LoginView + RegisterView", "Done", "M", 3, 0.5, "Built. UI-P6 toast → material.", "B001 Auth"),
    ("S003", "BiometricConsentView", "Done", "S", 1, 0.25, "biometric_consent_confirmed audit event wired.", "—"),
    ("S004", "VoiceExplanationView", "Done", "S", 1, 0.25, "Built.", "—"),
    ("S005", "VoiceRecordingView", "Done", "L", 8, 1.5, "Built (M-01 reconciled embedding dims). Raw audio purged immediately.", "—"),
    ("S006", "VoiceProcessingView", "Done", "M", 3, 0.5, "128-dim MFCC → Keychain only. Embedding never transmitted.", "S005"),
    ("S007", "WearableSetupView", "Done", "M", 3, 0.5, "iPhone fallback path; Ray-Ban Meta deferred.", "M019 Meta SDK (deferred)"),
    ("S008", "OnboardingFlowView", "Done", "M", 3, 0.5, "Step orchestration + accessibility passes.", "S003–S007"),
    ("S009", "PhysicianProfileSetupView", "Done", "M", 3, 0.5, "iPad-adaptive (UI-P6).", "B002 Profile"),
    ("S010", "MainTabView", "Done", "S", 1, 0.25, "UI-P2 native TabView + iPad sidebarAdaptable.", "S011, S013, S017"),
    ("S011", "DashboardView", "Done", "M", 3, 0.5, "M-07 Stage 2 tile; readable-measure clamp (UI-P6).", "B006, B010, M021"),
    ("S012", "CaptureView", "Done", "XL", 15, 3, "M-02/M-05/M-06 done. Live Activity start/pause/resume wired (UI-P4b).", "Apple Vision, AVFoundation, ReplayKit"),
    ("S013", "DeviceHubView", "Done", "M", 3, 0.5, "UI-P6 iPad clamp + a11y labels.", "M020 BLE"),
    ("S014", "PostEncounterView", "Done", "M", 3, 0.5, "Built.", "B010 Notes"),
    ("S015", "ProcessingView (in ContentView)", "Done", "L", 8, 1.5, "Stage 1 SLA + retry surface.", "B007, B010"),
    ("S016", "NoteReviewView", "Done", "XL", 15, 3, "M-09 source anchoring, M-10 conflicts, conflicts banner pulse (UI-P5), Stage 2 gate fix.", "B010, B011"),
    ("S017", "ExportView", "Done", "L", 8, 1.5, "On-device DOCX + Plain Text. PDF deferred.", "B012, M015"),
    ("S018", "NoteReadyView", "Done", "S", 1, 0.25, "Built.", "S014, S015"),
    ("S019", "SessionsInboxView", "Done", "M", 3, 0.5, "Search + filter + iPad clamp + value-based deep-link push.", "B006"),
    ("S020", "SessionNoteView", "Done", "M", 3, 0.5, "Spotlight NSUserActivity donation (UI-P4a).", "B010"),
    ("S021", "ProfileView", "Done", "M", 3, 0.5, "Voice profile mgmt + iPad-adaptive.", "B002"),
    ("S022", "FrameGalleryView (debug)", "Done", "S", 1, 0.25, "Hidden in release per #if DEBUG.", "—"),
    ("S023", "ContentView (root shell)", "Done", "M", 3, 0.5, "Root routing + scene-phase purge hook + onContinueUserActivity / onOpenURL.", "S001, S002, S008"),
    # Cross-cutting iOS UI / system work shipped this session
    ("UI-P1", "Color token sweep + semantic typography", "Done", "M", 3, 0.25, "Merged 2026-05-19 (commit 3a0152a).", "—"),
    ("UI-P2", "Native TabView + iPad sidebarAdaptable", "Done", "M", 3, 0.25, "Merged 2026-05-19 (commit 312d3fe).", "—"),
    ("UI-P3", "Inbox search + iPad clamp + amber conflicts banner", "Done", "L", 8, 0.5, "Merged 2026-05-19 (commit 40b3dcf).", "S019, S016"),
    ("UI-P4a", "App Intents + Spotlight integration", "Done", "M", 3, 0.5, "StartSessionIntent + ShowPendingNotesIntent + AppNavigation bus (commit df32abb).", "—"),
    ("UI-P4b", "Live Activity (Lock Screen + Dynamic Island) + Widget", "Done", "L", 8, 1, "AurionWidgets extension added via xcodeproj Ruby gem. StartSessionWidget + Live Activity widget (commit ba4900c + 9f3cca0).", "Apple Dev Program (now active)"),
    ("UI-P5", "Accessibility labels + symbol effects + motion polish", "Done", "M", 3, 0.25, "Capture controls, sort, toolbar, conflicts pulse (commit 4d00062).", "—"),
    ("UI-P6", "Materials (regularMaterial) + iPad readable-measure pass", "Done", "S", 1, 0.25, "Merged 2026-05-19 (commit 9cefee6).", "—"),
    ("AUR-DESIGN-NAVY", "Collapse aurionNavyLegacy → aurionNavy", "Done", "S", 0.5, 0.25, "Brand navy (#0C1B37) is canonical (commit d6a88d3).", "—"),
    ("AUR-DESIGN-DARK", "Full dark mode (muted slate palette)", "Done", "XL", 15, 1, "Adaptive tokens, 56 navy-text → adaptive swap, drop preferredColorScheme(.light) (commit 0cf99c8).", "AUR-DESIGN-NAVY"),
    ("Q-04", "SessionUIState shim cleanup", "Done", "S", 1, 0.25, "Drop Bool shims (commit 69b317e).", "—"),
    ("BUNDLE-ID", "Rebrand bundle ID → com.aurionclinical.aurion", "Done", "S", 0.5, 0.1, "Pre-TestFlight (commit 39c9404).", "—"),
    ("FASTLANE", "fastlane CLI bootstrap + TestFlight lanes", "Done", "M", 3, 0.25, "Appfile + Fastfile (commit e4a9d7b). User runs `fastlane bootstrap` once.", "Apple Dev active"),
    # Deferred mobile work
    ("M018", "MediaPipe face detection (backup)", "Future", "L", 5, 0.5, "Pilot follow-up — revisit only if clinical safety committee asks.", "M015 Apple Vision"),
    ("M019", "Ray-Ban Meta SDK", "Future", "XL", 15, 5, "Deferred post-pilot — pilot ships iPhone/iPad fallback.", "Partner SDK access"),
]

# ── WEB PORTAL — admin done, physician workflow greenfield ──────────────────
WEB = [
    # Admin portal — built & working
    ("W001", "Login / Auth (admin portal)", "Done", "S", 1, 0.25, "Real JWT cookie + role redirect.", "B001"),
    ("W002", "Dashboard — Pilot Metrics Overview", "Done", "M", 3, 0.5, "Real API + aggregated metrics.", "/admin/metrics, /admin/sessions"),
    ("W003", "Audit Log Viewer", "Done", "M", 3, 0.5, "DynamoDB stream + CSV export.", "Audit log populated"),
    ("W004", "PHI Masking Report", "Done", "M", 3, 0.5, "Real API + per-session breakdown.", "Comprehend Medical pipeline"),
    ("W005", "Provider Config Viewer", "Done", "S", 1, 0.25, "Read-only AppConfig fetch.", "AppConfig live"),
    ("W006", "Session List / Completeness", "Done", "M", 3, 0.5, "11-state badge map + specialty filter.", "/admin/sessions"),
    # Admin polish — verify after backend refactors
    ("W007", "User Management UI", "Verify", "L", 8, 1, "UI is final. Verify Cognito wiring after P0-06 (commit e7a5a90). Smoke test create/deactivate.", "WB14 Cognito AdminUser"),
    ("W008", "Eval Session List + Score Entry", "Verify", "L", 8, 1, "UI final. Verify EvalRecordModel persistence after B-08 refactor.", "WB13 EvalRecordModel"),
    # Admin missing pages
    ("W009", "Session Detail page", "Not Started", "M", 3, 0.5, "lib/api.ts has the getter; backend endpoint exists; page itself not built.", "WB12 (Done)"),
    # Deployment blockers
    ("W010", "Amplify Deployment + CORS Fix", "Blocker", "M", 5, 1, "Portal runs localhost:3000 only. No aws_amplify_app in TF. CORS still localhost-only.", "—"),
    ("W011", "Role-Based Navigation (Physician vs Admin)", "Not Started", "M", 4, 0.5, "Sidebar.tsx hardcoded to admin routes. Two-role guard pattern.", "Cognito role claim; W020 landing target"),
    # Physician workflow — data layer (foundation)
    ("W012", "TypeScript types (Patient/Appt/Visit/Schedule)", "Not Started", "S", 2, 0.5, "types/index.ts has 18 admin interfaces, zero physician.", "WB20–WB22 contracts frozen"),
    ("W013", "lib/api.ts physician API client", "Not Started", "S", 2, 0.5, "20 admin functions exist; zero physician.", "WB20–WB25 endpoint contracts"),
    ("W014", "WebSocket client (real-time Stage 1/2)", "Not Started", "M", 5, 1, "Backend /ws/{session_id} exists; no WS library in web package.json.", "W013, W028"),
    # Physician workflow — screens
    ("W020", "Schedule — Today's View (Day + Week)", "Not Started", "L", 10, 2, "Landing page for physicians. Appointment cards + status badges.", "WB21 Schedule API; W011 routing"),
    ("W021", "Patient Workspace shell + Overview tab", "Not Started", "L", 9, 2, "Demographics + latest visit + unresolved items.", "WB20, WB23"),
    ("W022", "Patient Workspace — Appointments tab", "Not Started", "M", 3, 0.5, "Reuses shell; read-only history.", "W021, WB21"),
    ("W023", "Patient Workspace — Visits tab", "Not Started", "M", 4, 0.75, "7 visit states + active drafts + approval state.", "W021, WB22"),
    ("W024", "Patient Workspace — Notes tab", "Not Started", "M", 4, 0.75, "Approved + draft + exported notes.", "W021, WB24, WB25"),
    ("W025", "Patient Workspace — Timeline tab", "Not Started", "M", 3, 0.5, "Longitudinal activity feed.", "W021, WB23"),
    ("W026", "Patient Workspace — Attachments tab", "Not Started", "S", 2, 0.5, "Minimal upload/list. No preview pipeline in MVP.", "W021, S3 upload endpoint"),
    ("W027", "Patient Workspace — Exports tab", "Not Started", "S", 2, 0.5, "Export history list.", "W021, WB25"),
    ("W028", "Visit Workspace / AI Documentation Review (CORE)", "Not Started", "XL", 24, 5, "CORE physician workflow. Real-time Stage 1/2 over WS + inline editing + approval. Largest single item.", "W014, WB22, WB24, WB25"),
    ("W029", "Conflict Resolution UI", "Not Started", "L", 11, 2, "ENRICHES/REPEATS/CONFLICTS panel.", "W028, WB24"),
    # Future
    ("W090", "French Localization", "Future", "L", 8, 3, "Deferred post-pilot. i18n scaffolding needed first.", "MVP strings externalised"),
    ("W091", "Patient-Facing Portal", "Future", "XL", 15, 6, "Out of MVP — new audience + auth surface.", "Patient identity/auth model"),
    ("W092", "Advanced Analytics Dashboard", "Future", "L", 8, 2, "Pilot dashboard adequate.", "Chart library selected"),
]

# ── BACKEND SERVICES — clinical mode complete, physician APIs pending ───────
BACKEND = [
    # Admin endpoints — Done
    ("B001", "Auth (api/v1/auth.py)", "Done", "M", 3, 0.5, "Cognito JWKS + JWT + dev seed (now @dataclass per Q-06).", "—"),
    ("B002", "Profile (api/v1/profile.py)", "Done", "S", 1, 0.25, "Physician profile CRUD.", "—"),
    ("B003", "Config (api/v1/config.py)", "Done", "S", 1, 0.25, "Runtime AppConfig.", "M008 AppConfig"),
    ("B004", "Admin (api/v1/admin.py — package split)", "Done", "L", 8, 1.5, "CQR-4 split into users/sessions/eval/metrics. Cognito wiring now real.", "B025 Alembic"),
    ("B005", "Privacy (api/v1/privacy.py)", "Done", "M", 3, 0.5, "PHI masking report + DSAR purge.", "B018"),
    ("B006", "Sessions (api/v1/sessions.py)", "Done", "L", 8, 1.5, "10-state lifecycle + paper-consent metadata.", "B021"),
    ("B007", "Transcription (api/v1/transcription.py)", "Done", "XL", 15, 2.5, "Whisper / AssemblyAI via registry + speaker tags + Stage 1 SLA.", "B015, M001"),
    ("B008", "Frames (api/v1/frames.py)", "Done", "M", 3, 0.5, "MaskingProof required on upload.", "B018"),
    ("B009", "Screen (api/v1/screen.py)", "Done", "L", 8, 1.5, "B-05 + M-08 vertical slice.", "B010, B019, M011"),
    ("B010", "Notes (api/v1/notes.py)", "Done", "XL", 15, 2.5, "Stage 1/2 + approve-stage1 async dispatch (M-09, M-10, B-06).", "B011, B022, B023"),
    ("B011", "Vision (api/v1/vision.py)", "Done", "XL", 15, 2.5, "Stage 2 captioning + conflict detection.", "B016 vision providers"),
    ("B012", "Export (api/v1/export.py)", "Done", "M", 3, 0.5, "On-device + audit endpoint (B-09).", "B020 Cleanup"),
    ("B013", "Health (api/v1/health.py)", "Done", "S", 1, 0.25, "Liveness for ECS/ALB.", "—"),
    ("B014", "WebSocket (api/v1/websocket.py)", "Done", "M", 3, 0.5, "Server-side delivered.", "B010"),
    ("B015", "Note-gen providers (modules/providers/note_gen/)", "Done", "XL", 15, 2.5, "OpenAI + Anthropic + Gemini.", "B024 registry"),
    ("B016", "Vision providers (modules/providers/vision/)", "Done", "L", 8, 1.5, "OpenAI + Anthropic + Gemini.", "B024 registry"),
    ("B017", "Transcription providers (modules/providers/transcription/)", "Done", "M", 3, 0.5, "Whisper + AssemblyAI.", "B024 registry"),
    ("B018", "Audit log service (modules/audit_log/)", "Done", "M", 3, 0.5, "DynamoDB append-only + AuditEventType StrEnum (Q-01) + kwarg whitelist (Q-03).", "M007 DynamoDB"),
    ("B019", "PHI audit service (modules/phi_audit/)", "Done", "M", 3, 0.5, "Comprehend Medical scan + audit-on-hit.", "M012 Comprehend Medical"),
    ("B020", "Cleanup service (modules/cleanup/)", "Done", "M", 3, 0.5, "Purge lifecycle + audited confirmation.", "M006 S3, B018"),
    ("B021", "Session state machine (modules/session/)", "Done", "L", 8, 1.5, "10-state + consent hard-block.", "B018"),
    ("B022", "Note versioning service (modules/note_gen/)", "Done", "L", 8, 1.5, "Immutable versions + approve marker + M-09 source anchoring.", "B010"),
    ("B023", "Stage 2 job queue (modules/vision/jobs.py)", "Done", "M", 3, 0.5, "Async pending→running→completed/failed (B-06).", "B011"),
    ("B024", "Provider registry (modules/config/provider_registry.py)", "Done", "M", 3, 0.5, "Type-safe; drives <30s AppConfig swap.", "M008"),
    ("B025", "Alembic migrations (P0-04)", "Done", "L", 8, 1.5, "Prod schema bootstrap + evolution (commit e330675).", "All models"),
    ("B026", "Persistent users + admin (P0-06)", "Done", "L", 8, 1.5, "_MOCK_USERS → SQL + Cognito (commit e7a5a90).", "B001, B025"),
    ("B027", "Persistent eval scoring (B-08)", "Done", "M", 3, 0.5, "EvalRecordModel + retention.", "B025"),
    ("B028", "End-to-end smoke test (P0-07)", "Done", "M", 3, 0.5, "in-process FastAPI + ASGITransport + SAVEPOINT rollback (commit 3835704).", "All modules"),
    # Cleanup ships
    ("Q-01", "AuditEventType StrEnum", "Done", "M", 2, 0.25, "Typo-safe audit event types (commit 5c26052).", "—"),
    ("Q-02", "privacy.py _purge_session_prefix extraction", "Done", "S", 1, 0.25, "Flattened purge nesting + metric count bugfix (commit ec3a318).", "—"),
    ("Q-03", "write_audit kwarg whitelist + strict mode", "Done", "S", 1, 0.25, "Runtime kwarg validator (commit d539a57).", "Q-01"),
    ("Q-05", "_to_uuid → core/uuids.py", "Done", "S", 0.5, 0.1, "4 duplicate copies → 1 helper (commit bdb22c3).", "—"),
    ("Q-06", "_DevUser → frozen @dataclass", "Done", "S", 0.5, 0.1, "Pydantic → frozen dataclass (commit 5c5f1e9).", "—"),
    # Physician workflow backend — Not Started (the big remaining backend slice)
    ("WB12", "GET /admin/sessions/{id} — Session Detail", "Done", "M", 3, 0.5, "Already covered by existing session detail endpoints; W009 page still missing.", "—"),
    ("WB20", "PatientModel + Patient CRUD API", "Not Started", "L", 10, 1.5, "Foundation for all physician workflow.", "B025 Alembic"),
    ("WB21", "AppointmentModel + Schedule API (Day/Week)", "Not Started", "L", 10, 1.5, "Powers W020.", "WB20"),
    ("WB22", "VisitModel + Visit CRUD + 7-state machine", "Not Started", "L", 11, 2, "Server-enforced 'one active visit' concurrency. CORE backend.", "WB20, WB21"),
    ("WB23", "Multi-Provider Continuity API", "Not Started", "M", 5, 0.75, "Nurse/resident observations feed.", "WB22"),
    ("WB24", "Note Approval + Conflict Gating API", "Not Started", "L", 10, 1.5, "Gate approve on unresolved CONFLICTS.", "WB22"),
    ("WB25", "Web-triggered Export API (DOCX + Plain Text)", "Not Started", "M", 5, 0.75, "Web parity with iOS export.", "WB22"),
    # Deferred
    ("B029", "Patient + chart (B-01, deferred to web)", "Future", "L", 8, 0, "Subsumed by WB20.", "—"),
    ("B030", "Schedule API (B-02, deferred to web)", "Future", "L", 8, 0, "Subsumed by WB21.", "—"),
    ("WB90", "EMR / FHIR Integration", "Future", "XL", 15, 6, "Post-pilot major workstream.", "FHIR resource mapping"),
    ("WB91", "Billing Workflow API", "Future", "L", 8, 3, "Post-pilot.", "Billing rules + payer integration"),
]

# ── INFRASTRUCTURE + AWS — production rollout phases tracked separately ─────
INFRA = [
    ("M001", "OpenAI Whisper", "Done", "M", 3, 0, "API via Secrets Manager.", "M013"),
    ("M002", "AssemblyAI", "Done", "M", 3, 0, "Selected via AppConfig.", "M013"),
    ("M003", "OpenAI GPT-4o (note gen + Stage 2)", "Done", "M", 3, 0, "Default per docker-compose.", "B015"),
    ("M004", "Anthropic Claude Sonnet 4", "Done", "M", 3, 0, "Note + vision.", "B015, B016"),
    ("M005", "Google Gemini 2.5 Flash", "Done", "M", 3, 0, "Eval comparisons; truncates long JSON.", "B015, B016"),
    ("M006", "AWS S3 (audio/frames/eval/export)", "Done", "M", 3, 0, "KMS-encrypted, public blocked, lifecycle.", "M010 KMS"),
    ("M007", "AWS DynamoDB (audit log)", "Done", "M", 3, 0, "Append-only + PITR + deletion protection.", "—"),
    ("M008", "AWS AppConfig", "Done", "M", 3, 0, "30s polling + schema validation.", "—"),
    ("M009", "AWS Cognito", "Done", "M", 3, 0, "JWT via JWKS.", "B026"),
    ("M010", "AWS KMS", "Done", "S", 1, 0, "All-services encryption.", "—"),
    ("M011", "AWS Textract (screen OCR)", "Done", "M", 3, 0, "On-device redact runs first.", "—"),
    ("M012", "AWS Comprehend Medical (PHI scan)", "Done", "M", 3, 0, "Audit on every hit.", "—"),
    ("M013", "AWS Secrets Manager", "Done", "S", 1, 0, "All AI provider keys.", "—"),
    ("M014", "PostgreSQL (RDS / docker)", "Done", "M", 3, 0, "Encrypted, 30d backup prod, multi-AZ var.", "B025 Alembic"),
    ("M015", "Apple Vision framework (face + OCR)", "Done", "L", 8, 0, "On-device fail-closed masking.", "—"),
    ("M016", "AVFoundation (capture)", "Done", "L", 8, 0, "Mode-aware audio/video.", "—"),
    ("M017", "ReplayKit (screen)", "Done", "M", 3, 0, "Feature-flagged.", "M015"),
    ("M020", "BLE pairing (BLEPairingManager)", "Done", "M", 3, 0, "Stubbed device list works.", "—"),
    ("M021", "WebSocket push (Backend ↔ iOS)", "Done", "M", 3, 0, "JWT-authenticated channel.", "B010"),
    ("M022", "iOS Keychain (auth token + voice embedding)", "Done", "S", 1, 0, "Embedding never transmits.", "—"),
    ("M023", "iOS UserDefaults (crash recovery)", "Done", "S", 1, 0, "M-05 consent metadata persists.", "—"),
    ("M024", "LocalStack (dev infra)", "Done", "M", 3, 0, "S3/DDB/AppConfig/Cognito seed script.", "—"),
    ("M025", "Terraform/HCL (Phase 8)", "Done", "L", 8, 0, "13 .tf files in ca-central-1; remote state + dev/prod tfvars.", "M006–M013"),
    # AWS Amplify — blocker for web portal
    ("WM10", "AWS Amplify — web portal hosting", "Blocker", "M", 4, 1, "No aws_amplify_app in TF. Bundled into Prod Phase 4 work.", "W010"),
]

# ── PRODUCTION ROLLOUT — Phases 1-5 ─────────────────────────────────────────
PROD = [
    ("P1", "Phase 1 — State bootstrap (S3 + DDB + KMS)", "Done", "S", 0.5, 0.25, "bootstrap/ module applied to dev (commit cd40d3d). Account 366034225426.", "—"),
    ("P2", "Phase 2 — TLS + DNS + WAF + SHA-pin", "Done", "M", 1.5, 1, "Fully applied 2026-05-22 (115 resources). api-dev.aurionclinical.com live, ACM cert valid, WAF in front, ALB returning HTTPS. ECS service waiting on first image push.", "P1"),
    ("P3", "Phase 3 — GitHub Actions CI/CD with OIDC", "Verify", "M", 1, 0.5, "Workflows + IAM OIDC trust shipped (commit 2a0fd11). Roles exist in AWS. 4 GitHub Secrets need wiring: AWS_ACCOUNT_ID, AWS_DEPLOY_ROLE_DEV, AWS_DEPLOY_ROLE_PROD, ECR_REGISTRY.", "P1, P2"),
    ("P4", "Phase 4 — Cognito MFA + SNS + alarms + CloudTrail + Flow Logs + KMS rotation", "Not Started", "M", 1.5, 1, "Includes 5 missing operational alarms, SNS topic, GuardDuty enable.", "P2"),
    ("P5", "Phase 5 — Runbooks + IAM Identity Center + backup-restore drill", "Not Started", "M", 1, 1.5, "Mostly markdown + 1 IAM rework. Pre-pilot human-gated.", "P4"),
]

# ── EXTERNAL BLOCKERS — non-engineering, can't be compressed ────────────────
EXTERNAL = [
    ("EXT-01", "Apple Developer Program activation", "Done", "—", 0, 0, "Active 2026-05-22. Unblocks TestFlight, App IDs, Live Activity push (if added).", "—"),
    ("EXT-02", "Cloudflare DNS delegation (subdomain to AWS)", "Done", "—", 0, 0, "Option B chosen + applied 2026-05-22. 4 NS records at Cloudflare for api-dev.aurionclinical.com delegate to AWS. Propagation verified via dig.", "—"),
    ("EXT-03", "First Phase 2 full apply against dev", "Done", "—", 0, 0, "Applied 2026-05-22. 115 resources created. ALB returns 503 (no image yet — expected). TLS valid.", "EXT-02"),
    ("EXT-04", "TestFlight initial build review", "Not Started", "—", 0, 0, "Apple reviews first build before internal testers can install. 24-48h.", "fastlane bootstrap; first iOS push"),
    ("EXT-05", "Cognito MFA enrollment by pilot physicians", "Not Started", "—", 0, 0, "~5 min per physician via Cognito hosted UI.", "P4 (MFA enable)"),
    ("EXT-06", "DPIA + consent forms + clinical safety committee sign-off", "Blocker", "—", 0, 0, "NON-ENGINEERING. 2-4 weeks of calendar depending on pilot site's IRB/QI committee cadence.", "—"),
    ("EXT-07", "Cyber insurance + Business Associate Agreements (if applicable)", "Blocker", "—", 0, 0, "NON-ENGINEERING. Weeks. Required before live PHI.", "—"),
    ("EXT-08", "Pilot consent + recruitment of 3-5 physicians", "Not Started", "—", 0, 0, "Dr. Perry + Dr. Marie confirmed. Other 1-3 pending.", "—"),
]


# ─────────────────────────────────────────────────────────────────────────────
# Workbook building
# ─────────────────────────────────────────────────────────────────────────────

def write_inventory_sheet(ws, title: str, lane: str, rows: list):
    """Render an inventory sheet with the standard column set."""
    headers = [
        "ID", "Name / Description", "Status", "S/M/L/XL",
        "Human-Days (P50)", "Claude-Code-Days", "Notes / Commit", "Pre-requisites",
    ]
    # Title row
    ws.cell(1, 1, value=title).font = Font(bold=True, size=14)
    ws.cell(2, 1, value=f"Lane: {lane} · status as-of 2026-05-22").font = Font(italic=True, color="595959")
    # Header row
    for col, h in enumerate(headers, start=1):
        c = ws.cell(4, col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP_CENTER
        c.border = ALL_BORDER

    # Data rows
    for i, row in enumerate(rows, start=5):
        rid, name, status, complexity, human_d, claude_d, notes, prereq = row
        ws.cell(i, 1, value=rid).alignment = WRAP
        ws.cell(i, 2, value=name).alignment = WRAP
        emoji = STATUS_EMOJI.get(status, "")
        ws.cell(i, 3, value=f"{emoji} {status}").alignment = WRAP_CENTER
        ws.cell(i, 4, value=complexity).alignment = WRAP_CENTER
        ws.cell(i, 5, value=human_d).alignment = WRAP_CENTER
        ws.cell(i, 6, value=claude_d).alignment = WRAP_CENTER
        ws.cell(i, 7, value=notes).alignment = WRAP
        ws.cell(i, 8, value=prereq).alignment = WRAP
        # Fill per status
        fill = STATUS_FILLS.get(status)
        if fill:
            for col in range(1, 9):
                ws.cell(i, col).fill = fill
                ws.cell(i, col).border = ALL_BORDER

    # Totals row
    total_row = 5 + len(rows) + 1
    ws.cell(total_row, 2, value="TOTALS").font = Font(bold=True)
    # Sum Human-Days
    ws.cell(total_row, 5, value=f"=SUM(E5:E{4 + len(rows)})").font = Font(bold=True)
    ws.cell(total_row, 6, value=f"=SUM(F5:F{4 + len(rows)})").font = Font(bold=True)
    # Remaining (not-done) sums
    rem_row = total_row + 1
    ws.cell(rem_row, 2, value="REMAINING (not 🟢 Done, excl. 🟣 Future)").font = Font(bold=True, italic=True)
    ws.cell(rem_row, 5, value=(
        f'=SUMPRODUCT((NOT(ISNUMBER(SEARCH("Done", C5:C{4 + len(rows)})))) * '
        f'(NOT(ISNUMBER(SEARCH("Future", C5:C{4 + len(rows)})))) * E5:E{4 + len(rows)})'
    )).font = Font(bold=True)
    ws.cell(rem_row, 6, value=(
        f'=SUMPRODUCT((NOT(ISNUMBER(SEARCH("Done", C5:C{4 + len(rows)})))) * '
        f'(NOT(ISNUMBER(SEARCH("Future", C5:C{4 + len(rows)})))) * F5:F{4 + len(rows)})'
    )).font = Font(bold=True)

    # Column widths
    widths = [16, 50, 16, 9, 16, 17, 60, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"


def write_overview_sheet(ws):
    """Top-level summary that pulls from each inventory sheet."""
    ws.cell(1, 1, value="AURION — Effort Estimation v5 (Consolidated)").font = Font(bold=True, size=16)
    ws.cell(2, 1, value=(
        "Mobile + Web Portal + Backend + Infra + Production Rollout. "
        "Reconciled 2026-05-22. Claude-Code-Days reflects the autonomous "
        "workflow pace (AURION-CODING-WORKFLOW.md), not human-dev pace."
    )).alignment = WRAP
    ws.row_dimensions[2].height = 50

    # Status palette explainer
    ws.cell(4, 1, value="LEGEND").font = Font(bold=True)
    legend = [
        ("🟢 Done", "Shipped + verified. Commit SHA in Notes column."),
        ("🟡 Verify", "Code shipped; needs end-to-end verification before claiming Done."),
        ("🟠 Not Started", "MVP scope. 0% built."),
        ("🔴 Blocker", "Hard prerequisite for the next phase. Resolve first."),
        ("🔵 In Progress", "Partial — some sub-tasks shipped, others pending."),
        ("🟣 Future", "Out-of-MVP. Estimated for visibility but not in pilot path."),
    ]
    for i, (k, v) in enumerate(legend, start=5):
        ws.cell(i, 1, value=k).alignment = WRAP_CENTER
        ws.cell(i, 2, value=v).alignment = WRAP

    # Roll-up table
    row = 13
    ws.cell(row, 1, value="ROLL-UP BY LANE").font = Font(bold=True, size=13)
    row += 2
    headers = ["Lane", "Items", "Done", "Verify", "Not Started", "Blocker", "Future", "Human-Days Remaining", "Claude-Days Remaining"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row, col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP_CENTER
        c.border = ALL_BORDER
    row += 1

    # Each lane references its inventory sheet
    lanes = [
        ("Mobile (iOS)",          "Mobile",          len(MOBILE)),
        ("Web Portal",            "Web Portal",      len(WEB)),
        ("Backend Services",      "Backend",         len(BACKEND)),
        ("Infrastructure / AWS",  "Infrastructure",  len(INFRA)),
        ("Production Rollout",    "Production Rollout", len(PROD)),
        ("External Blockers",     "External",        len(EXTERNAL)),
    ]
    for label, sheet, n in lanes:
        last = 4 + n   # data rows 5..(4+n)
        ws.cell(row, 1, value=label)
        ws.cell(row, 2, value=n)
        ws.cell(row, 3, value=f'=COUNTIF(\'{sheet}\'!C5:C{last},"*Done*")')
        ws.cell(row, 4, value=f'=COUNTIF(\'{sheet}\'!C5:C{last},"*Verify*")')
        ws.cell(row, 5, value=f'=COUNTIF(\'{sheet}\'!C5:C{last},"*Not Started*")')
        ws.cell(row, 6, value=f'=COUNTIF(\'{sheet}\'!C5:C{last},"*Blocker*")')
        ws.cell(row, 7, value=f'=COUNTIF(\'{sheet}\'!C5:C{last},"*Future*")')
        # Remaining = sum of Human/Claude days where status is NOT Done and NOT Future
        ws.cell(row, 8, value=(
            f'=SUMPRODUCT((NOT(ISNUMBER(SEARCH("Done",\'{sheet}\'!C5:C{last}))))*'
            f'(NOT(ISNUMBER(SEARCH("Future",\'{sheet}\'!C5:C{last}))))*\'{sheet}\'!E5:E{last})'
        ))
        ws.cell(row, 9, value=(
            f'=SUMPRODUCT((NOT(ISNUMBER(SEARCH("Done",\'{sheet}\'!C5:C{last}))))*'
            f'(NOT(ISNUMBER(SEARCH("Future",\'{sheet}\'!C5:C{last}))))*\'{sheet}\'!F5:F{last})'
        ))
        for col in range(1, 10):
            ws.cell(row, col).border = ALL_BORDER
        row += 1

    # Grand totals
    ws.cell(row, 1, value="GRAND TOTAL").font = Font(bold=True)
    ws.cell(row, 2, value=f"=SUM(B{row - len(lanes)}:B{row - 1})").font = Font(bold=True)
    for col in range(3, 10):
        ws.cell(row, col, value=f"=SUM({get_column_letter(col)}{row - len(lanes)}:{get_column_letter(col)}{row - 1})").font = Font(bold=True)
    for col in range(1, 10):
        ws.cell(row, col).fill = PatternFill("solid", fgColor="D9E1F2")
        ws.cell(row, col).border = ALL_BORDER

    # Critical-path narrative
    row += 3
    ws.cell(row, 1, value="CRITICAL PATH TO PILOT").font = Font(bold=True, size=13)
    row += 1
    cp = [
        ("Backend physician workflow (WB20 → WB21 → WB22 → WB24, serial)", "~6 days"),
        ("Web frontend physician workflow (W021 → W028 → W029, after backend ready)", "~8 days"),
        ("Web admin polish + Amplify + Prod Phase 4-5 (parallel)", "~3 days"),
        ("Tests + go-live gate", "~2 days"),
        ("iOS polish + Config.swift dev/prod swap", "~1 day"),
        ("CRITICAL PATH TOTAL (with lane parallelism)", "~13 working days"),
        ("", ""),
        ("EXTERNAL BLOCKERS (calendar, NOT engineering)", ""),
        ("Apple Dev activation", "DONE 2026-05-22"),
        ("Domain delegation (Cloudflare) + Phase 2 apply", "~1 day"),
        ("Initial Phase 2 full apply + smoke test", "~1 day"),
        ("Pilot physician onboarding + TestFlight invites", "~2-3 days"),
        ("DPIA + consent forms + clinical safety committee sign-off", "~2-4 weeks"),
        ("", ""),
        ("Pilot-ready calendar estimate", "5-7 weeks from 2026-05-22 → 2026-06-26 to 2026-07-10"),
        ("Demo target", "2026-07-15 (~1 week buffer)"),
    ]
    for label, val in cp:
        c1 = ws.cell(row, 1, value=label)
        c2 = ws.cell(row, 2, value=val)
        if "CRITICAL PATH TOTAL" in label or "Pilot-ready calendar" in label or "Demo target" in label:
            c1.font = Font(bold=True)
            c2.font = Font(bold=True)
        if "EXTERNAL BLOCKERS" in label:
            c1.font = Font(bold=True, color="C00000")
        row += 1

    # Conversion-rate note
    row += 1
    ws.cell(row, 1, value="CLAUDE-CODE-PACE CONVERSION (observed in this build session)").font = Font(bold=True)
    row += 1
    notes = [
        ("S (1d human) → 0.1–0.25 Claude-d", "Single-file edits, small refactors, config swaps."),
        ("M (3d human) → 0.5–0.75 Claude-d", "Multi-file refactors, new endpoints, screen scaffolds."),
        ("L (8d human) → 1.0–2.0 Claude-d", "Cross-cutting features, new modules, integration work."),
        ("XL (15-24d human) → 2.5–5.0 Claude-d", "Net-new screens with state machine + tests (W028 Visit Workspace is the upper bound)."),
        ("External blockers (Apple, DNS, IRB, insurance)", "No compression. Wall-clock dominated."),
    ]
    for label, val in notes:
        ws.cell(row, 1, value=label)
        ws.cell(row, 2, value=val).alignment = WRAP
        row += 1

    # Column widths for overview
    for col, w in enumerate([55, 60, 14, 14, 14, 14, 14, 22, 22], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


def main():
    wb = Workbook()
    # First sheet — Overview
    overview = wb.active
    overview.title = "Overview"
    write_overview_sheet(overview)

    sheets = [
        ("Mobile",             "MOBILE — iOS App",                   "Mobile (iOS)",    MOBILE),
        ("Web Portal",         "WEB PORTAL — Admin + Physician",     "Web (Next.js)",   WEB),
        ("Backend",            "BACKEND SERVICES",                   "FastAPI",         BACKEND),
        ("Infrastructure",     "INFRASTRUCTURE & AWS",               "Terraform/AWS",   INFRA),
        ("Production Rollout", "PRODUCTION ROLLOUT (5 phases)",      "Terraform/CI",    PROD),
        ("External",           "EXTERNAL BLOCKERS (non-engineering)", "External",       EXTERNAL),
    ]
    for sheet_name, title, lane, rows in sheets:
        ws = wb.create_sheet(sheet_name)
        write_inventory_sheet(ws, title, lane, rows)

    wb.save(OUT_PATH)
    print(f"wrote {OUT_PATH}")
    print(f"  Mobile:             {len(MOBILE)} rows")
    print(f"  Web Portal:         {len(WEB)} rows")
    print(f"  Backend:            {len(BACKEND)} rows")
    print(f"  Infrastructure:     {len(INFRA)} rows")
    print(f"  Production Rollout: {len(PROD)} rows")
    print(f"  External:           {len(EXTERNAL)} rows")


if __name__ == "__main__":
    main()
